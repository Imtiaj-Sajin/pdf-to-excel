# ══════════════════════════════════════════════════════════════════════════
# LEASE PDF EXTRACTOR v5.1 — FIXED ADDENDUM EXTRACTION
# ══════════════════════════════════════════════════════════════════════════
!pip install pdfplumber openpyxl -q

import pdfplumber, re, os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from google.colab import files

# ─── Constants ─────────────────────────────────────────────────────────────
MONTH_NAMES = {
    'january','february','march','april','may','june','july',
    'august','september','october','november','december'
}
ORDINAL_RE = re.compile(r'^\d{1,2}(st|nd|rd|th)$', re.I)
YEAR_RE    = re.compile(r'^(20\d{2}|19\d{2})$')
AMOUNT_RE  = re.compile(r'^\d{1,6}\.\d{2}$')

# ─── Core helpers ──────────────────────────────────────────────────────────
def despace_numbers(text):
    for _ in range(8):
        prev = text
        text = re.sub(r'(\d) (\d)', r'\1\2', text)
        text = re.sub(r'(\d) (\.)',  r'\1\2', text)
        text = re.sub(r'(\.) (\d)',  r'\1\2', text)
        if text == prev: break
    return text

def ordinal_suffix(n):
    n = int(n)
    if 11 <= (n % 100) <= 13: return f"{n}th"
    return f"{n}{['th','st','nd','rd','th'][min(n%10,4)]}"

def crop_text(page, x0, top, x1, bottom):
    try:
        c = page.crop((x0, top, min(x1, page.width), min(bottom, page.height)))
        return c.extract_text(x_tolerance=3, y_tolerance=3) or ''
    except Exception:
        return ''

def upright_words(page):
    return sorted(
        [w for w in page.extract_words(x_tolerance=3, y_tolerance=3)
         if w.get('upright', True) and w['height'] > 2],
        key=lambda w: (round(w['top']), w['x0'])
    )

# ══════════════════════════════════════════════════════════════════════════
# GROUP 1 — MAIN LEASE PAGE  (unchanged from v5)
# ══════════════════════════════════════════════════════════════════════════
PARTIES_SKIP = {
    'lease','contract','list','all','people','signing','the','is','between',
    'you','resident','residents','sometimes','referred','to','as','this','a',
    'an','and','of','in','for','or','1.','2.','3.','4.','parties','parties.',
    'name','names','cont','tenant','landlord','llc','mf','park','aston',
    'southern','breeze','davenport','florida','floor','unit','apt',
}

def _is_name_word(text):
    t = text.strip(' ,.')
    if len(t) < 2: return False
    if t.lower() in PARTIES_SKIP: return False
    if re.search(r'\d', t): return False
    if re.search(r'[(){}[\]<>\\/@#%^&*+=|~`:;!?"]', t): return False
    if not t[0].isupper(): return False
    for word in t.split():
        if len(word) > 1 and not word.isupper():
            if any(c.isupper() for c in word[1:]):
                return False
    return True

def _extract_month_from_contract_garble(word):
    template = list("contract")
    remaining = []
    t_idx = 0
    for c in word.lower():
        if t_idx < len(template) and c == template[t_idx]: t_idx += 1
        else: remaining.append(c)
    candidate = ''.join(remaining)
    if candidate.lower() in MONTH_NAMES:
        return candidate[0].upper() + candidate[1:].lower()
    return None

def get_date_of_contract(pages):
    page  = pages[0]
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    for w in words:
        if re.search(r'Date\s+of\s+Lease\s+Contract', w['text'], re.I):
            ly, lx1 = w['top'], w['x1']
            candidates = [nw for nw in words
                          if abs(nw['top'] - ly) < 18 and nw['x0'] > lx1 - 5
                          and len(nw['text'].strip()) > 1]
            for c in candidates:
                if (re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)',
                              c['text'], re.I) and re.search(r'\d{4}', c['text'])):
                    return c['text'].strip()
            for c in candidates:
                if c['text'].strip().lower() in MONTH_NAMES:
                    row = sorted([nw for nw in words if abs(nw['top']-c['top'])<5],
                                 key=lambda x: x['x0'])
                    parts = [rw['text'].strip(',') for rw in row
                             if rw['text'].strip(',').lower() in MONTH_NAMES
                             or re.match(r'\d{1,2},?$', rw['text'].strip())
                             or YEAR_RE.match(rw['text'].strip())]
                    if parts: return ' '.join(parts)
    t = crop_text(page, 0, 50, page.width, 100)
    m = re.search(r'\b([A-Za-z]+ \d{1,2},?\s*\d{4})\b', t)
    return m.group(1).strip() if m else "Not found"

def get_parties(pages):
    page  = pages[0]
    pw    = page.width
    LEFT  = pw * 0.52
    all_w = upright_words(page)
    anchor_y = None
    for w in all_w:
        if w['x0'] >= LEFT: continue
        if re.search(r'lease\s*contract\s*\)?[:\s]*$', w['text'], re.I):
            anchor_y = w['bottom']; break
    if anchor_y is None: anchor_y = 148.0
    owner_y = None
    for w in all_w:
        if w['x0'] >= LEFT or w['top'] < anchor_y: continue
        if re.search(r'\band\s+us\b', w['text'], re.I):
            owner_y = w['top']; break
        if re.search(r'\bowner\b', w['text'], re.I) and w['top'] > anchor_y + 60:
            owner_y = w['top']; break
    if owner_y is None: owner_y = anchor_y + 145
    name_parts = []
    for w in all_w:
        if w['x0'] >= LEFT or w['top'] < anchor_y - 2: continue
        if w['top'] >= owner_y: break
        if _is_name_word(w['text']): name_parts.append(w['text'].strip())
    if name_parts:
        result = re.sub(r'\s+', ' ', ' '.join(name_parts)).strip(' ,')
        result = re.sub(r'\s*(Aston|Park|LLC|MF)\b.*$', '', result, flags=re.I).strip(' ,')
        if result: return result
    return "Not found"

def get_lease_term(pages):
    for pg in pages:
        words = sorted(
            [w for w in pg.extract_words(x_tolerance=3, y_tolerance=3)
             if w.get('upright', True) and w['height'] > 2 and w['x0'] > 310],
            key=lambda w: (round(w['top']), w['x0'])
        )
        ordinals = [w for w in words if ORDINAL_RE.match(w['text'].strip())]
        months   = [w for w in words if w['text'].strip().lower() in MONTH_NAMES]
        years    = [w for w in words if YEAR_RE.match(w['text'].strip())]
        begin = end = None

        if ordinals and months and years:
            begin = (f"{months[0]['text'].rstrip(',').strip()} "
                     f"{ordinals[0]['text']}, {years[0]['text']}")
            if len(ordinals)>=2 and len(months)>=2 and len(years)>=2:
                end = (f"{months[-1]['text'].rstrip(',').strip()} "
                       f"{ordinals[-1]['text']}, {years[-1]['text']}")

        if begin and not end:
            t = despace_numbers(crop_text(pg, 318, 130, pg.width, 200))
            parts = re.split(r'11:59\s*p\.?m\.?', t, maxsplit=1, flags=re.I)
            if len(parts) > 1:
                m = re.search(r'the\s+(.{2,15}?)\s+day\s+of\s+(.{3,25}?)\s*,\s*(\d{4})',
                              parts[1], re.I|re.S)
                if m:
                    d,mo,yr = re.sub(r'\s','',m.group(1)),re.sub(r'\s','',m.group(2)),m.group(3).strip()
                    if ORDINAL_RE.match(d) and mo.lower() in MONTH_NAMES:
                        end = f"{mo[0].upper()+mo[1:].lower()} {d}, {yr}"

        if begin and not end:
            edate_w = sorted([w for w in words if 156<w['top']<170 and w['height']>=9],
                             key=lambda w: w['x0'])
            if edate_w:
                ord_digits = ''
                for w in edate_w:
                    if 320<w['x0']<395: ord_digits=''.join(c for c in w['text'] if c.isdigit()); break
                month_str = None
                for w in edate_w:
                    if 425<w['x0']<500:
                        clean=re.sub(r'\s','',w['text'])
                        month_str=(clean[0].upper()+clean[1:].lower() if clean.lower() in MONTH_NAMES
                                   else _extract_month_from_contract_garble(clean))
                        if month_str: break
                yr_digits=''.join(c for w in edate_w if w['x0']>505 for c in w['text'] if c.isdigit())
                if ord_digits and month_str and len(yr_digits)>=4:
                    end = f"{month_str} {ordinal_suffix(ord_digits)}, {yr_digits[:4]}"

        if not begin or not end:
            t = despace_numbers(crop_text(pg, 318, 130, pg.width, 200))
            if not begin:
                m = re.search(r'begins\s+on\s+the\s+(.{2,8}?)\s+day\s+of\s+(.{3,15}?)\s*,\s*(\d{4})',
                              t, re.I|re.S)
                if m:
                    d,mo,yr=re.sub(r'\s','',m.group(1)),re.sub(r'\s','',m.group(2)),m.group(3).strip()
                    if ORDINAL_RE.match(d) and mo.lower() in MONTH_NAMES:
                        begin=f"{mo[0].upper()+mo[1:].lower()} {d}, {yr}"
            if not end:
                parts = re.split(r'11:59\s*p\.?m\.?', t, maxsplit=1, flags=re.I)
                if len(parts)>1:
                    m=re.search(r'the\s+(.{2,15}?)\s+day\s+of\s+(.{3,25}?)\s*,\s*(\d{4})',
                                parts[1],re.I|re.S)
                    if m:
                        d,mo,yr=re.sub(r'\s','',m.group(1)),re.sub(r'\s','',m.group(2)),m.group(3).strip()
                        if ORDINAL_RE.match(d) and mo.lower() in MONTH_NAMES:
                            end=f"{mo[0].upper()+mo[1:].lower()} {d}, {yr}"

        if begin or end: return begin or "Not found", end or "Not found"
    return "Not found", "Not found"

def get_security_deposit(pages):
    for pg in pages:
        all_w = pg.extract_words(x_tolerance=3, y_tolerance=3)
        right_w = [w for w in all_w if w.get('upright',True) and w['x0']>310 and 560<w['top']<660]
        for w in sorted(right_w, key=lambda x: x['top']):
            val = w['text'].replace(',','').lstrip('$').strip()
            if AMOUNT_RE.match(val): return val
        for y0,y1 in [(560,660),(540,680)]:
            t = despace_numbers(crop_text(pg, 318, y0, pg.width, y1))
            m = re.search(r'residents?\s+in\s+the\s+apartment\s+is\s+\$\s*([\d.]+)', t, re.I)
            if m and AMOUNT_RE.match(m.group(1)): return m.group(1)
            if re.search(r'deposit|residents', t, re.I):
                for found in re.findall(r'\$\s*(\d+\.\d{2})', t):
                    if AMOUNT_RE.match(found): return found
        garble_w = sorted(
            [w for w in all_w if w.get('upright',True) and w['height']>=9
             and 430<w['x0']<530 and 612<w['top']<634],
            key=lambda w: w['x0'])
        if garble_w:
            raw = ''.join(c for w in garble_w for c in w['text'] if c.isdigit() or c=='.')
            m = re.search(r'\d{2,6}\.\d{2}', raw)
            if m and AMOUNT_RE.match(m.group()): return m.group()
    return "Not found"

def get_monthly_rent(pages):
    for pg in pages:
        all_w = pg.extract_words(x_tolerance=3, y_tolerance=3)
        full_t = despace_numbers(pg.extract_text(x_tolerance=3, y_tolerance=3) or '')
        for w in all_w:
            if w.get('upright',True) and w['text'].lower().strip('.,')=='month':
                row=[nw for nw in all_w if abs(nw['top']-w['top'])<10]
                row_t=' '.join(nw['text'].lower() for nw in sorted(row,key=lambda x:x['x0']))
                if 'per' in row_t:
                    for rw in sorted(row,key=lambda x:x['x0']):
                        val=rw['text'].replace(',','').lstrip('$').strip()
                        if AMOUNT_RE.match(val): return val
        for pat in [
            r'RENT\s+AND\s+CHARGES.*?you\s+will\s+pay\s*\$?\s*([\d.]+)\s*per\s+month',
            r'\$\s*([\d.]+)\s+per\s+month\s+for\s+rent',
            r'pay\s*\$?\s*([\d.]+)\s*per\s+month',
        ]:
            m=re.search(pat,full_t,re.I|re.S)
            if m and AMOUNT_RE.match(m.group(1)): return m.group(1)
        for w in sorted(all_w,key=lambda x:x['top']):
            if AMOUNT_RE.match(w['text'].strip()):
                if 45<w['x0']<200 and 470<w['top']<525: return w['text'].strip()
        m=re.search(r'sum\s+of\s+\$\s*([\d.]+)\s*\(equal',full_t,re.I)
        if m and AMOUNT_RE.match(m.group(1)): return m.group(1)
        liq_w=[w for w in all_w if w.get('upright',True) and w['x0']>310
               and 255<w['top']<290 and AMOUNT_RE.match(w['text'].replace(',','').lstrip('$').strip())]
        if liq_w: return liq_w[0]['text'].replace(',','').lstrip('$').strip()
    return "Not found"

# ══════════════════════════════════════════════════════════════════════════
# GROUP 2 — RENT CONCESSION ADDENDUM  (completely rewritten)
# ══════════════════════════════════════════════════════════════════════════

def find_addendum_page(pdf):
    """Scan ALL pages for LEASE ADDENDUM FOR RENT CONCESSION."""
    for pg in pdf.pages:
        # Try extract_text first
        t = pg.extract_text(x_tolerance=3, y_tolerance=3) or ''
        if re.search(r'LEASE\s+ADDENDUM\s+FOR\s+RENT\s+CONCES', t, re.I):
            return pg
        # Fallback: word-level scan (handles garbled text pages)
        for w in pg.extract_words(x_tolerance=3, y_tolerance=3):
            if re.search(r'ADDENDUM.*RENT|RENT.*CONCES', w['text'], re.I):
                return pg
    return None

def _safe_amount(raw):
    """Extract first valid XX.XX amount from a raw string."""
    raw = despace_numbers(str(raw).replace(',', ''))
    for tok in raw.split():
        tok = tok.lstrip('$').strip('.')
        if AMOUNT_RE.match(tok):
            return tok
    m = re.search(r'\d{1,6}\.\d{2}', raw)
    return m.group() if m else ""

def get_addendum_data(page):
    """
    Extract all 5 concession fields using extract_text + regex.
    This avoids the word-merging issue from extract_words.
    """
    if page is None:
        return "", "", "", "", ""

    pw, ph = page.width, page.height

    # ── Find section 3 start y ─────────────────────────────────
    sec3_y = ph * 0.45   # safe default ~453pt
    for w in page.extract_words(x_tolerance=3, y_tolerance=3):
        if re.search(r'CONCESSION.{0,5}DISCOUNT.{0,5}AGREEMENT', w['text'], re.I):
            sec3_y = max(0, w['top'] - 15)
            break

    # ── Crop LEFT column only, from section 3 downward ─────────
    left = page.crop((0, sec3_y, pw * 0.52, ph))
    raw  = left.extract_text(x_tolerance=3, y_tolerance=3) or ''
    text = despace_numbers(raw)

    # ── 1. One-Time Concession amount ──────────────────────────
    # Pattern: "total amount of $  14325.00  ."
    onetime_amt = ""
    m = re.search(r'total\s+amount\s+of\s+\$\s*([\d\s,]+\.?\d*)', text, re.I)
    if m:
        onetime_amt = _safe_amount(m.group(1))

    # Word-level fallback with tight tolerance (avoids merging)
    if not onetime_amt:
        tight = sorted(
            [w for w in page.extract_words(x_tolerance=1, y_tolerance=2)
             if w.get('upright', True) and w['height'] > 2
             and w['x0'] < pw * 0.52 and w['top'] > sec3_y],
            key=lambda w: (round(w['top']), w['x0']))
        for i, w in enumerate(tight):
            if re.search(r'total\s+amount', w['text'], re.I):
                for nw in tight[i:i+10]:
                    v = _safe_amount(nw['text'])
                    if v: onetime_amt = v; break
                break

    # ── 2. Concession months (text on blank lines after "month(s) of:") ──
    onetime_months = ""
    m = re.search(
        r'for\s+the\s+month\(?s\)?\s+of\s*:\s*(.*?)(?=\n\s*\n\s*\n|\$\s*\n|Monthly\s+Discount|Other\s+Discount)',
        text, re.I | re.S)
    if not m:
        # Broader fallback
        m = re.search(r'for\s+the\s+month\(?s\)?\s+of\s*:\s*(.*?)(?:Monthly\s+Discount|Other\s+Discount|\Z)',
                      text, re.I | re.S)
    if m:
        raw_m = re.sub(r'\s+', ' ', m.group(1)).strip()
        # Strip trailing period/dot then clean
        raw_m = raw_m.rstrip('. ')
        if raw_m and len(raw_m) > 2:
            # Remove trailing ". " artifacts
            raw_m = re.sub(r'\s*\.\s*$', '', raw_m).strip()
            onetime_months = raw_m

    # ── 3. Monthly Discount amount ──────────────────────────────
    monthly_disc = ""
    m = re.search(r'Monthly\s+Discount\s+of\s+\$\s*([\d\s,\.]+?)(?:\s+per\s+month|\s*\n|\Z)',
                  text, re.I)
    if m:
        monthly_disc = _safe_amount(m.group(1))

    # ── 4. Other Discount amount ────────────────────────────────
    other_amt = ""
    # Matches "Lease Contract: $  295.00" or "Contract: $ 295.00"
    m = re.search(r'[Cc]ontract\s*:\s*\$\s*([\d\s,\.]+?)(?:\s*\.|\s+Reason|\s*\n|\Z)',
                  text, re.I)
    if m:
        other_amt = _safe_amount(m.group(1))

    # ── 5. Other Discount comment (lines after "below:") ────────
    other_note = ""
    m = re.search(r'below\s*:\s*(.*?)(?:Resident\s+or|Owner\s+or|\Z)',
                  text, re.I | re.S)
    if m:
        note = re.sub(r'\s+', ' ', m.group(1)).strip()
        note = note.strip('. ')
        if note and len(note) > 2:
            other_note = note

    return onetime_amt, onetime_months, monthly_disc, other_amt, other_note

# ══════════════════════════════════════════════════════════════════════════
# ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════

def process_pdf(pdf_path):
    fname = os.path.basename(pdf_path)
    print(f"\n{'─'*62}\n  📄  {fname}\n{'─'*62}")

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        n     = min(4, total)
        pages = [pdf.pages[i] for i in range(n)]
        print(f"  Pages: {total}  |  Scanning first {n} for main data\n")

        date       = get_date_of_contract(pages)
        parties    = get_parties(pages)
        begin, end = get_lease_term(pages)
        deposit    = get_security_deposit(pages)
        rent       = get_monthly_rent(pages)

        add_page = find_addendum_page(pdf)
        if add_page:
            print(f"  📋  Addendum found → page {add_page.page_number} of {total}")
        else:
            print("  ⚠️   Addendum NOT found")

        onetime_amt, onetime_months, monthly_disc, other_amt, other_note = \
            get_addendum_data(add_page)

    row = {
        'File Name':              fname,
        'Date of Lease Contract': date,
        'Parties (Residents)':    parties,
        'Lease Begin Date':       begin,
        'Lease End Date':         end,
        'Security Deposit ($)':   deposit,
        'Monthly Rent ($)':       rent,
        ' ':                      '',
        'One-Time Concession ($)': onetime_amt,
        'Concession Month(s)':     onetime_months,
        'Monthly Discount ($)':    monthly_disc,
        'Other Discount ($)':      other_amt,
        'Other Discount Comment':  other_note,
    }

    G1 = ['File Name','Date of Lease Contract','Parties (Residents)',
          'Lease Begin Date','Lease End Date','Security Deposit ($)','Monthly Rent ($)']
    G2 = ['One-Time Concession ($)','Concession Month(s)',
          'Monthly Discount ($)','Other Discount ($)','Other Discount Comment']
    pad = max(len(k) for k in row if k.strip())

    print(f"\n  {'─'*20} MAIN LEASE {'─'*22}")
    for k in G1:
        v = row[k]
        icon = '✅' if v not in ('Not found','') else '⚠️ '
        print(f"  {icon}  {k:<{pad}}  →  {v}")

    print(f"\n  {'─'*18} CONCESSION ADDENDUM {'─'*18}")
    for k in G2:
        v = row[k]
        icon = '✅' if v not in ('','Not found') else '—  '
        print(f"  {icon}  {k:<{pad}}  →  {v if v else '(blank — not filled in PDF)'}")

    return row

# ══════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════
COL_G1  = ['File Name','Date of Lease Contract','Parties (Residents)',
            'Lease Begin Date','Lease End Date','Security Deposit ($)','Monthly Rent ($)']
COL_GAP = [' ']
COL_G2  = ['One-Time Concession ($)','Concession Month(s)',
            'Monthly Discount ($)','Other Discount ($)','Other Discount Comment']

HDR_G1 = PatternFill('solid', fgColor='1F4E79')
HDR_G2 = PatternFill('solid', fgColor='2E75B6')
HDR_GP = PatternFill('solid', fgColor='D9D9D9')
FNT_W  = Font(bold=True, color='FFFFFF', size=11)
FNT_G  = Font(bold=True, color='808080', size=11)

def to_excel(rows, path='lease_data_extracted.xlsx'):
    col_order = COL_G1 + COL_GAP + COL_G2
    df = pd.DataFrame(rows, columns=col_order)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Lease Data')
        ws = writer.sheets['Lease Data']
        g1s, g2s = set(COL_G1), set(COL_G2)
        for cell in ws[1]:
            hdr = cell.value or ''
            cell.fill = HDR_G1 if hdr in g1s else HDR_G2 if hdr in g2s else HDR_GP
            cell.font = FNT_G if hdr.strip() == '' else FNT_W
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 34
        for col in ws.columns:
            hdr = (col[0].value or '').strip()
            ws.column_dimensions[col[0].column_letter].width = (
                3 if not hdr else min(max(len(str(c.value or '')) for c in col)+4, 55))
        ws.freeze_panes = 'A2'
    print(f"\n  ✅  Excel → {path}")
    return path

# ══════════════════════════════════════════════════════════════════════════
# RUN
# ══════════════════════════════════════════════════════════════════════════
print("📁  Upload lease PDF files…")
uploaded = files.upload()

rows = []
for fname, content in uploaded.items():
    tmp = f'/tmp/{fname}'
    with open(tmp, 'wb') as f: f.write(content)
    try:
        rows.append(process_pdf(tmp))
    except Exception:
        import traceback; traceback.print_exc()
        rows.append({'File Name': fname,
                     **{k:'ERROR' for k in COL_G1[1:]+COL_G2}, ' ':''})

if rows:
    files.download(to_excel(rows))
    print("\n🎉  Done!")
